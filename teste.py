from openpyxl import Workbook

# Create workbook and sheets
wb = Workbook()

# Seccionadora sheet
ws_sec = wb.active
ws_sec.title = "Seccionadora"
ws_sec.append(["RELATÓRIO – SECCIONADORA"])
ws_sec.append([])
ws_sec.append(["Marca", ""])
ws_sec.append(["Modelo", ""])
ws_sec.append(["Espessura da serra (mm)", ""])
ws_sec.append(["Tipo (Automática/Manual)", ""])
ws_sec.append(["Refilo no comprimento (mm)", ""])
ws_sec.append(["Refilo na largura (mm)", ""])
ws_sec.append(["Tamanho mínimo da peça - Comprimento (mm)", ""])
ws_sec.append(["Tamanho mínimo da peça - Largura (mm)", ""])
ws_sec.append(["Nº de chapas empilhadas", ""])
ws_sec.append(["Tipo de impressora (Nenhuma/Manual/Automática)", ""])
ws_sec.append(["Modelo da impressora (se manual)", ""])
ws_sec.append(["Observações", ""])

# Coladeira sheet
ws_col = wb.create_sheet(title="Coladeira")
ws_col.append(["RELATÓRIO – COLADEIRA DE BORDO"])
ws_col.append([])
ws_col.append(["Marca", ""])
ws_col.append(["Modelo", ""])
ws_col.append(["Sobra de fita entre peças (mm)", ""])
ws_col.append(["Limpeza da tupia de entrada (mm)", ""])
ws_col.append(["Tamanho mínimo da peça - Comprimento (mm)", ""])
ws_col.append(["Tamanho mínimo da peça - Largura (mm)", ""])
ws_col.append(["Largura mínima da peça (mm)", ""])
ws_col.append(["Espessuras de fita utilizadas (mm)", ""])
ws_col.append(["Espessuras de chapa utilizadas (mm)", ""])
ws_col.append(["Observações", ""])

# Centro de Usinagem sheet
ws_centro = wb.create_sheet(title="Centro de Usinagem")
ws_centro.append(["RELATÓRIO – CENTRO DE FURAÇÃO / USINAGEM"])
ws_centro.append([])
ws_centro.append(["Marca", ""])
ws_centro.append(["Modelo", ""])
ws_centro.append(["Comprimento mínimo da peça (mm)", ""])
ws_centro.append(["Largura mínima da peça (mm)", ""])
ws_centro.append(["Comprimento máximo da peça (mm)", ""])
ws_centro.append(["Largura máxima da peça (mm)", ""])
ws_centro.append(["Observações", ""])

# Save file
file_path = "Relatorio_Maquinas_Marcenaria.xlsx"
wb.save(file_path)

file_path
